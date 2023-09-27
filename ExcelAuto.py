import openpyxl
import pyclip
import time
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)

def get_file_path():
    return input("Please enter the path of the Excel file: ")

def get_section():
    return input("Section: (e.g. A1, B1): ")

def get_rsNum():
    return input("Please enter the specific RS number here: ")

def get_condition():
    return input("Please enter the name of the condition: ")

excel_path = get_file_path()
specific_sheet = get_condition().upper()
rs_number = get_rsNum().upper()
section = get_section().upper()

if get_condition() == "A":
    specific_sheet = "25C"

if get_condition() == "H":
    specific_sheet = "50C"

if get_condition() == "R":
    specific_sheet = "0C"

def copy_clipboard(content):
    copied_data = pyclip.paste() 

    if copied_data:
        
        # "/Users/youngsong/Desktop/ExcelAutoTest.xlsx"
        
        workbook = openpyxl.load_workbook(excel_path)
        print("available sheets: ", workbook.sheetnames)
        
        matching_sheets = [sheet for sheet in workbook.sheetnames if specific_sheet == sheet]

        if matching_sheets:
            worksheet = workbook[matching_sheets[0]]
        else:
            print("Not Found")
            return 
    
        #Identify target column
        target_column = None
        section_column = None

        print(f"file: {excel_path}")
        print(f"sheet: {specific_sheet}")
        print(f"rs number: {rs_number}")
        print(f"section: {section}")

        for col in worksheet.iter_cols(min_row=3, max_row=4):
            rs_cell = col[0]
            section_cell = col[1]
            print(f"rs cell: {rs_cell.value} (Formula: {rs_cell.formula})")
            print(f"sectoion cell: {section_cell.value} (Formula: {section_cell.formula})")
            
            if (rs_cell.value == rs_number or rs_cell.formula == rs_number) and (section_cell.value == section or section_cell.formula == section):
                target_column = rs_cell.column
                break

        #Paste clipboard data
        if target_column:   
            data_rows = copied_data.strip().split("\n")
            for row_number, data in enumerate(data_rows, start=5):
                worksheet.cell(row = row_number, column= target_column, value= data)
                workbook.save(excel_path)
                print("Data successfully pasted into Excel file!")
        else:
            print(f"Tag {rs_number} not found")
    else:
        print("No data found in the clipboard")


last_clipboard_content = pyclip.paste()

while True:
    clipboard_content = pyclip.paste()
    if clipboard_content != last_clipboard_content:
        copy_clipboard(clipboard_content)
        last_clipboard_content = clipboard_content
        break
    
    time.sleep(1)